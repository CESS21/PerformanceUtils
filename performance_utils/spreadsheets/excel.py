# -*- coding: utf-8 -*-
"""
Microsoft Office Excel spreadsheet utilities.

This module provides spreadsheet utilities for Microsoft Office Excel.
"""

import openpyxl

from typing import List
import os

import performance_utils.spreadsheets.common as common


class Worksheet(common.AbstractCRUDWorksheet):
    """
    Excel worksheet.

    Attributes:
        name: Title of the worksheet.
        _IR: Internal representation of worksheet.
        _workbook: Workbook object the worksheet is contained in.
    """

    name: str
    _IR: openpyxl.worksheet.worksheet.Worksheet
    _workbook: "Workbook"

    def __init__(
        self,
        workbook: "Workbook",
        name: str,
        index: int = None
        ) -> None:
        """
        Initializes `Worksheet` object.

        Args:
            workbook: Workbook to open/create sheet in.
            name: Title of the worksheet.
            index: Position to insert new sheet in list. If None, appends to
                sheet list. Unnecessary if sheet already exists in workbook.
        """

        super().__init__(workbook, name, index)

        try:
            self._IR = self._workbook._IR[self.name]
        except KeyError:
            if index is None:
                self._IR = self._workbook._IR.create_sheet(self.name)
            else:
                self._IR = self._workbook._IR.create_sheet(self.name, index)

    def delete(self) -> None:
        """Destroys worksheet."""

        # Delete from actual workbook.
        del self._workbook._IR[self.name]
        # Delete from list of sheets.
        for (index, worksheet) in enumerate(self._workbook._sheets):
            if worksheet.name == self.name:
                del self._workbook._sheets[index]
    
    def read(self, row: int, column: int) -> str:
        """
        Retrieves the content of a worksheet cell.

        Args:
            row: Index of the row to retrieve cell content from.
            column: Index of the column to retrieve cell content from.

        Returns:
            Cell contents.
        """

        return str(self._IR.cell(row, column).value)

    def update(self, row: int, column: int, data: str) -> None:
        """
        Updates the content of a worksheet cell.

        Args:
            row: Index of the row to update cell content in.
            column: Index of the column to update cell content in.
            data: Data to update the cell content with.
        """

        self._IR.cell(row, column, data)


class Workbook(common.AbstractCRUDWorkbook):
    """
    Excel workbook.

    Use `with` keyword to create and use instances.

    Attributes:
        name: Title of the Excel workbook.
        _sheets: List of worksheets within the workbook.
        _IR: Internal representation of workbook.
    """

    name: str
    _sheets: List[Worksheet]
    _IR: openpyxl.workbook.workbook.Workbook

    def __init__(self, name: str) -> None:
        """
        Initialized `Workbook` object.

        Args:
            name: Title of the workbook.
        """

        super().__init__(name)

    def __enter__(self):
        """Called upon `with` block entry."""

        super().__enter__()
        return self

    def __exit__(self, exc_type, exc_value, traceback):
        """Called upon `with` block exit."""

        super().__exit__(exc_type, exc_value, traceback)

    def create(self) -> None:
        """Opens an Excel workbook or creates a new one."""

        try:
            self._IR = openpyxl.load_workbook("{0}.xlsx".format(self.name))
        except OSError: # File does not exist.
            self._IR = openpyxl.Workbook()
            self.save()
        finally:
            for sheetname in self._IR.sheetnames:
                self._sheets.append(Worksheet(self, sheetname))

    def delete(self) -> None:
        """Destroys workbook."""

        os.remove("{0}.xlsx".format(self.name))

    def sheet(self, sheet_name: str, index: int = None) -> Worksheet:
        """
        Opens a worksheet by name if it exists or creates a new one.

        Args:
            sheet_name: Title of worksheet.
            index: Position to insert new sheet in list. If None, appends to
                sheet list. Unnecessary if sheet already exists.

        Returns:
            Worksheet wrapped in `Worksheet` object.
        """

        for worksheet in self._sheets:
            if worksheet.name == sheet_name:
                return worksheet

        # If worksheet was not found...
        if index is None:
            worksheet = Worksheet(self, sheet_name)
            self._sheets.append(worksheet)
        else:
            worksheet = Worksheet(self, sheet_name, index)
            self._sheets.insert(index, worksheet)
        
        return worksheet

    def save(self) -> None:
        """Saves the workbook to storage."""

        self._IR.save("{0}.xlsx".format(self.name))
